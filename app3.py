import pandas as pd
import streamlit as st
import lxml
import urllib.request
from bokeh.models.widgets import Button
from bokeh.models import CustomJS
from streamlit_bokeh_events import streamlit_bokeh_events
from streamlit_lottie import st_lottie
import requests
import streamlit.components.v1 as components
import json
import openpyxl

st.set_page_config(page_title="AgroAppCredicoop",page_icon="",layout="wide") 

#unpacking
workbook = openpyxl.load_workbook(r'C:\Users\Usuario\Desktop\dataframe.xlsx')
worksheet = workbook.active
header = [cell.value for cell in next(worksheet.iter_rows())]
data = [cell.value for row in worksheet.iter_rows(min_row=2) for cell in row]
result = dict(zip(header, data))
for key, value in result.items():
    globals()[key] = value

def copy_button():
    copy_button = Button(label="Copiar tabla")
    copy_button.js_on_event("button_click", CustomJS(args=dict(df=st.session_state.dfa.to_csv(sep='\t')), code="""
        navigator.clipboard.writeText(df);
        """))
    no_event = streamlit_bokeh_events(
        copy_button,
        events="GET_TEXT",
        key="get_text",
        refresh_on_update=True,
        override_height=75,
        debounce_time=0)
    
def css():
    # CSS to inject contained in a string
    hide_table_row_index = """
            <style>
            thead tr th:first-child {display:none}
            tbody th {display:none}
            </style>
            """
    # Inject CSS with Markdown
    st.markdown(hide_table_row_index, unsafe_allow_html=True)

def load_lottieurl(url: str):
    r = requests.get(url)
    if r.status_code != 200:
        return None
    return r.json()

def app():
    st.title(" Valuaci贸n de hacienda")
    left, right = st.columns(2)
    left.write("Completar:")
    form = left.form("template_form")
    tipo = form.selectbox('Ingrese tipo de hacienda: ', ["Ternero             ", "Novillito       ", "Ternera             ", "Vaquillona        ", "Vaca                "])
    cantidad = form.number_input("Ingrese cantidad de cabezas: ", step=1)
    peso = form.number_input("Ingrese peso: ", step=1)
    submit = form.form_submit_button("Ingresar")
    df=pd.read_html('https://www.monasterio-tattersall.com/precios-hacienda') #leo la tabla de la p谩gina
    hacienda = df[0] 
    categoria = hacienda.Categor铆a 
    promedio = hacienda.Promedio
    tabla = pd.DataFrame({'categoria':categoria,'promedio':promedio}) #creo un dataframe con categoria y promedio
    ternero=tabla[0:4] 
    novillito=tabla[4:7]
    ternera=tabla[7:11]
    vaquillona=tabla[11:14]
    vaca=tabla[19:20]
    fecha=(tabla[25:26].values)[0][0]
    ternero160=int(ternero.promedio[0][2:5])
    ternero180=int(ternero.promedio[1][2:5])
    ternero200=int(ternero.promedio[2][2:5])
    ternero230=int(ternero.promedio[3][2:5])
    novillo260=int(novillito.promedio[4][2:5])
    novillo300=int(novillito.promedio[5][2:5])
    novillo301=int(novillito.promedio[6][2:5])
    ternera150=int(ternera.promedio[7][2:5])
    ternera170=int(ternera.promedio[8][2:5])
    ternera190=int(ternera.promedio[9][2:5])
    ternera210=int(ternera.promedio[10][2:5])
    vaquillona250=int(vaquillona.promedio[11][2:5])
    vaquillona290=int(vaquillona.promedio[12][2:5])
    vaquillona291=int(vaquillona.promedio[13][2:5])
    vacas=int(vaca.promedio[19][2:8])
    def constructor():
        def valores():
            if tipo == 'Ternero             ' and peso < 160:
                valor = ternero160*cantidad*peso
            elif tipo == 'Ternero             ' and peso < 180:
                valor = ternero180*cantidad*peso
            elif tipo == 'Ternero             ' and peso <= 200:
                valor = ternero200*cantidad*peso
            elif tipo == 'Ternero             ' and peso > 200:
                valor = ternero230*cantidad*peso
            elif tipo == 'Ternero             ' and peso == 0:
                valor = ternero200*cantidad*peso
            elif tipo == 'Novillito       ' and peso < 260:
                valor = novillo260*cantidad*peso
            elif tipo == 'Novillito       ' and peso <= 300:
                valor = novillo300*cantidad*peso
            elif tipo == 'Novillito       ' and peso > 300:
                valor = novillo301*cantidad*peso
            elif tipo == 'Novillito       ' and peso == 0:
                valor = novillo300*cantidad*peso
            elif tipo == 'Ternera             ' and peso < 150:
                valor = ternera150*cantidad*peso
            elif tipo == 'Ternera             ' and peso < 170:
                valor = ternera170*cantidad*peso
            elif tipo == 'Ternera             ' and peso <= 190:
                valor = ternera190*cantidad*peso
            elif tipo == 'Ternera             ' and peso > 190:
                valor = ternera210*cantidad*peso
            elif tipo == 'Ternera             ' and peso == 0:
                valor = ternera190*cantidad*peso
            elif tipo == 'Vaquillona        ' and peso < 250:
                valor = vaquillona250*cantidad*peso
            elif tipo == 'Vaquillona        ' and peso <= 290:
                valor = vaquillona290*cantidad*peso
            elif tipo == 'Vaquillona        ' and peso > 290:
                valor = vaquillona291*cantidad*peso
            elif tipo == 'Vaquillona        ' and peso == 0:
                valor = vaquillona290*cantidad*peso
            elif tipo == 'Vaca                ':
                valor = vacas*cantidad
            valor = int(valor*0.9)
            return valor #valor de ajuste
        valor=valores()
        d = [tipo, cantidad, peso, valor]
        return d
    metalista=[]
    if "dfa" not in st.session_state:
        st.session_state.dfa = pd.DataFrame(columns=("Categor铆a", "Cantidad", "Peso", "Valuaci贸n"))
    if submit:
        metalista.append(constructor())
        dfb = pd.DataFrame(metalista, columns=("Categor铆a", "Cantidad", "Peso", "Valuaci贸n"))
        st.session_state.dfa = pd.concat([st.session_state.dfa, dfb])
    css()
    valuacion_total = st.session_state.dfa['Valuaci贸n'].sum()
    right.metric('La valuaci贸n total de hacienda es: ', '${:,}'.format(valuacion_total))
    right.write("Tabla para copiar:")
    right.table(st.session_state.dfa.style.format({"Cantidad":"{:.0f}", "Peso":"{:.0f}", "Valuaci贸n":"${:,}"}))
    right.write(f'Los precios considerados son de la {fecha}')
    promedios = pd.DataFrame(
        {'Categoria': ['Ternero', 'Novillo', 'Ternera', 'Vaquillonas'],
         'Peso': ['180', '260', '170','250']})
    st.write(f'Pesos promedio para tipo de hacienda (en caso que no se informe el peso). En vacas poner peso cero')
    st.table(promedios.assign(hack='').set_index('hack'))
    
def app1():
    df2=pd.read_html('https://www.cotagroweb.com.ar/pizarra/')
    data2 = df2[0]
    # psoja= 58760 en caso que falle precio cotagro habilitar esta l铆nea
    psoja = data2.iloc[0,1]
    ppsoja = int(psoja[1:])
    pmaiz= data2.iloc[1,1]
    ppmaiz = int(pmaiz[1:])
    ptrigo= data2.iloc[2,1]
    pptrigo = int(ptrigo[1:])
    pgira= data2.iloc[4,1]
    ppgira = 74400 #int(pgira[1:])
    ppsorgo = 44400
    fecha = data2.columns[1][7:]
    st.title(" Valuaci贸n de granos")
    st.write(f'Precios de pizarra del Mercado de Rosario al {fecha}')
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Soja", '${:,}'.format(int(ppsoja)))
    col2.metric("Trigo", '${:,}'.format(int(pptrigo)))
    col3.metric("Ma铆z", '${:,}'.format(int(ppmaiz)))
    col4.metric("Sorgo", '${:,}'.format(int(ppsorgo)))
    col5.metric("Girasol",'${:,}'.format(int(ppgira)))
    left, right = st.columns(2)
    left.write("Completar:")
    form = left.form("template_form")
    tipo = form.selectbox('Ingrese tipo de grano: ', ["Soja","Trigo","Ma铆z","Sorgo","Girasol"])
    cantidad = form.number_input("Ingrese toneladas: ", step=1)
    submit = form.form_submit_button("Ingresar")
    def lista():
        def valor():
            if tipo == "Soja":
                precio = ppsoja
            elif tipo == "Trigo":
                precio = pptrigo
            elif tipo == "Ma铆z":
                precio = ppmaiz
            elif tipo == "Sorgo":
                precio = ppsorgo
            else:
                precio = ppgira
            return int(cantidad*precio)
        valor = valor()
        lista = [tipo, cantidad, valor]
        return lista
    cereales=[]
    if "dfs" not in st.session_state:
        st.session_state.dfs = pd.DataFrame(columns=("Tipo grano", "Cantidad (tn)", "Valuaci贸n"))
    if submit:
        cereales.append(lista())
        dfd = pd.DataFrame(cereales, columns=("Tipo grano", "Cantidad (tn)", "Valuaci贸n"))
        st.session_state.dfs = pd.concat([st.session_state.dfs, dfd])
    css()
    valuacion_total = st.session_state.dfs['Valuaci贸n'].sum()
    right.metric('La valuaci贸n total de granos es: ', '${:,}'.format(valuacion_total))
    right.write("Tabla para copiar:")
    right.table(st.session_state.dfs.style.format({"Cantidad (tn)":"{:.0f}", "Valuaci贸n":"${:,}"}))

def app2():
    if "ingresos_totales" not in st.session_state:
        st.session_state["ingresos_totales"] = 0
    st.title(" Servicios agr铆colas")
    left, right = st.columns(2)
    left.write("Completar:")
    form = left.form("template_form")
    tipo = form.selectbox('Ingrese tipo de servicio: ', ["Cosecha","Siembra","Pulverizaci贸n","Laboreos"])
    cantidad = form.number_input("Ingrese superficie (has): ", step=1)
    precio = form.number_input("Ingrese precio por ha", step=1)
    submit = form.form_submit_button("Ingresar")
    valorminc = 9000 #valor minimo cosecha
    valormaxc = 16000 #valor maximo cosecha
    valors = 7500 #valor referencia siembra
    valormins = valors*0.50 #valor minimo siembra
    valormaxs = valors*1.50 #valor maximo siembra
    
    def lista():
        def valor():
            return cantidad*precio
        valor = valor()
        lista = [tipo, cantidad, precio, valor]
        return lista
    servagro=[]
    if "dfx" not in st.session_state:
        st.session_state.dfx = pd.DataFrame(columns=("Categor铆a", "Superficie(ha)", "Precio", "Ingreso estimado"))
    if submit:
        servagro.append(lista())
        st.session_state["ingresos_totales"] += cantidad*precio
        dfy = pd.DataFrame(servagro, columns=("Categor铆a", "Superficie(ha)", "Precio", "Ingreso estimado"))
        st.session_state.dfx = pd.concat([st.session_state.dfx, dfy])
        if tipo == 'Cosecha' and (precio > valormaxc or precio < valorminc):
            st.warning("ALERTA! El precio por ha de cosecha cargado es fuera de los promedios de mercado. Ver precios de referencia abajo")
        elif tipo == 'Siembra' and (precio > valormaxs or precio < valormins):
            st.warning("ALERTA! El precio por ha de siembra cargado es fuera de los promedios de mercado. Ver precios de referencia abajo")
        else:
            pass
    css()
    right.metric('Los ingresos totales por servicios agr铆colas son: ', "${:,}".format(st.session_state["ingresos_totales"]))
    right.write("Tabla para copiar:")
    right.table(st.session_state.dfx.style.format({"Superficie(ha)":"{:.0f}", "Precio":"${:,}", "Ingreso estimado":"${:,}"}))
    
    
    def mostrar_precios_referencia(tipo_servicio, imagen):
        expander = st.expander(f"Ver precios de referencia - {tipo_servicio}")
        expander.image(imagen)
    mostrar_precios_referencia("Cosecha Soja", "https://www.agrocontratistas.com.ar/img/Precios/SOJA202303.jpg")
    mostrar_precios_referencia("Cosecha Ma铆z", "https://www.agrocontratistas.com.ar/img/Precios/MAIZ202303.jpg")
    mostrar_precios_referencia("Siembra y Laboreos", "https://www.agrocontratistas.com.ar/img/Precios/Labores_202302.jpg")
    return st.session_state.dfx
    
def app3():
    st.title("锔 Estado de los campos")
    with st.expander("Recomendaciones de interpretaci贸n"):
     st.write("""
         - Para ver el panorama general de sequ铆a ir a 驴Qu茅 zonas estan en sequ铆a? y buscar en "unidad administrativa de nivel 2" la localidad donde estan los campos
         - En caso de estar en 谩rea de sequ铆a ver la secci贸n "Evoluci贸n de sequ铆as entre dos per铆odos" para ver si se registraron mejoras en los ultimos meses.
         - En la secci贸n 驴Hace cuanto que no llueve? se puede ver la 煤ltima informaci贸n de precipitaciones
         - Tener en cuenta que el mapa de calor se conforma con la informaci贸n recolectada de las estaciones por lo que algunas 谩reas con pocas estaciones (como por ejemplo zona centro este de Santa Fe) pueden verse influenciadas por estaciones m谩s lejanas
     """)
    components.iframe("https://dashboard.crc-sas.org/informes/como-estamos/", height = 1500)
    st.caption("Datos extraidos de https://sissa.crc-sas.org/novedades/publicaciones-y-reportes-tecnicos/")
    

def app4():
    st.title(" Planteo productivo")
    
    # Obtener el valor actual de tipo de cultivo en la sesi贸n actual
    tipo_cultivo = st.session_state.get('tipo_cultivo', None)

    # Si no hay un valor para tipo de cultivo, establecer el valor predeterminado como el primer elemento de la lista
    if tipo_cultivo is None:
        tipo_cultivo = "Soja 1ra"
    
    # Crear un men煤 desplegable para el tipo de cultivo y actualizar el valor en la sesi贸n actual
    tipo_cultivo = st.selectbox('Tipo de cultivo: ', ["Soja 1ra", "Soja 2da", "Trigo","Ma铆z","Girasol", "Sorgo", "Cebada"], index=[i for i, x in enumerate(["Soja 1ra", "Soja 2da", "Trigo","Ma铆z","Girasol", "Sorgo", "Cebada"]) if x == tipo_cultivo][0])
    st.session_state['tipo_cultivo'] = tipo_cultivo
    
    left, right = st.columns(2)
    left.write("Completar:")
    form = left.form("template_form")

    # Cerrar el formulario
    form.empty()

    # Si se ha seleccionado un tipo de cultivo, mostrar un men煤 desplegable para las regiones
    if tipo_cultivo:
        # Obtener el valor actual de la regi贸n en la sesi贸n actual
        region = st.session_state.get('region', None)

        # Si no hay un valor para la regi贸n, establecer el valor predeterminado como el primer elemento de la lista correspondiente al tipo de cultivo
        if region is None:
            if tipo_cultivo == "Soja 1ra":
                region = "Regi贸n 1"
            elif tipo_cultivo == "Soja 2da":
                region = "Regi贸n 4"

            st.session_state['region'] = region

        # Crear un men煤 desplegable para la regi贸n y actualizar el valor en la sesi贸n actual
        region_options = []
        if tipo_cultivo == "Soja 1ra":
            region_options = ["Regi贸n 1", "Regi贸n 2", "Regi贸n 3"]
        elif tipo_cultivo == "Soja 2da":
            region_options = ["Regi贸n 4", "Regi贸n 5", "Regi贸n 6"]
        elif tipo_cultivo == "Trigo":
            region_options = ["Regi贸n 7", "Regi贸n 8", "Regi贸n 9"]
        elif tipo_cultivo == "Ma铆z":
            region_options = ["Regi贸n 10", "Regi贸n 11", "Regi贸n 12"]
        elif tipo_cultivo == "Girasol":
            region_options = ["Regi贸n 13", "Regi贸n 14", "Regi贸n 15"]
        elif tipo_cultivo == "Sorgo":
            region_options = ["Regi贸n 16", "Regi贸n 17", "Regi贸n 18"]
        elif tipo_cultivo == "Cebada":
            region_options = ["Regi贸n 19", "Regi贸n 20", "Regi贸n 21"]

    # Crear un men煤 desplegable para la regi贸n
    region = form.selectbox('Regi贸n: ', region_options)
    propio = form.selectbox('Tipo de explotaci贸n: ', ["Propia","Arrendado","Aparcer铆a"])
    cantidad = form.number_input("Superficie (has): ", step=1)
    rinde = form.number_input("Rendimiento informado (en tn)")

    # ... c贸digo del formulario ...
    datos = []
    if form.form_submit_button("Ingresar"):
        # Agregar las variables a la lista
        datos.append([region, propio, cantidad])
        df = pd.DataFrame(datos, columns=['Regi贸n', 'Tipo de explotaci贸n', 'Superficie (has)'])
        # Imprimir la lista de datos
        st.table(df)
        css()
        
        # API tipo de cambio
    url = "https://www.dolarsi.com/api/api.php?type=valoresprincipales"
    response = requests.get(url)
    if response.status_code == 200:
        api_data = response.json()
        value = api_data[0]['casa']['venta']
        value2 = value.replace(',', '.')
        dol = float(value2)
    else:
        print("Failed to retrieve data")
        
    right.metric("Dolar oficial", '${:,}'.format(float(dol)))
    right.write("Cuadro gastos:")
    gastos = right.number_input("Gastos de estructura", step=1)
    arrendamiento = right.number_input("Gastos de arrendamiento", step=1)
    aparceria = right.number_input("Porcentaje de aparcer铆a", step=1)
    
    #precio = psoja1*dol*rinde*cantidad
    
    #costos directos
    #costodirecto = costo*dol*cantidad
    
    #gasto de comercializaci贸n
    #gastoscom = porgastos*ibsoja1
    
def app5():
     st.header("Cuadro resumen")
     left, right = st.columns(2)
     css()
     left.subheader(" Existencias de granos")
     left.table(st.session_state.dfs.style.format({"Cantidad (tn)":"{:.0f}", "Valuaci贸n":"${:,}"}))
     left.subheader(" Ingresos Servicios agr铆colas")
     left.table(st.session_state.dfx.style.format({"Superficie(ha)":"{:.0f}", "Precio":"${:,}", "Ingreso estimado":"${:,}"}))
     right.subheader(" Existencias de hacienda")
     right.table(st.session_state.dfa.style.format({"Cantidad":"{:.0f}", "Peso":"{:.0f}", "Valuaci贸n":"${:,}"}))
     
#configuraciones de p谩gina   
lottie_book = load_lottieurl('https://assets7.lottiefiles.com/packages/lf20_d7OjnJ.json')
with st.sidebar:
    st.title('Agro App')
    st.markdown("---")
my_button = st.sidebar.radio("Modulos",('Planteo productivo', 'Condiciones clim谩ticas', 'Valuaci贸n granos', 'Valuaci贸n hacienda', 'Servicios agr铆colas', 'Cuadro resumen'))
if my_button == 'Valuaci贸n hacienda':
    app()
elif my_button == 'Valuaci贸n granos':
    app1()
elif my_button == 'Servicios agr铆colas':
    app2()
elif my_button == 'Condiciones clim谩ticas':
    app3()
elif my_button == 'Cuadro resumen':
    app5()
else:    
    app4()
with st.sidebar:
    st.markdown("---")
    st_lottie(lottie_book, speed=0.5, height=50, key="initial")
    st.markdown("---")
    st.caption("Desarrollado por JSantacecilia para Equipo Agro Banco Credicoop")
    
    
# Mantenimiento app
# psorgo l149
# parametros servicios agricolas l210 - 211 -212 y expanders l256, 258 y 260

#hacienda st.session_state.dfa
#granos st.session_state.dfs
#servicios st.session_state.dfx